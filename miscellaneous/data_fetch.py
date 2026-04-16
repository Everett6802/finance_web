#! /usr/bin/python
# -*- coding: utf8 -*-

import os
import sys
import re
import argparse
import errno
# '''
# Question: How to Solve xlrd.biffh.XLRDError: Excel xlsx file; not supported ?
# Answer : The latest version of xlrd(2.01) only supports .xls files. Installing the older version 1.2.0 to open .xlsx files.
# '''
# import xlrd
# import xlsxwriter
from openpyxl import Workbook, load_workbook
import math
import pandas as pd
import csv
import argparse
from datetime import datetime, date, timedelta
import getpass
from collections import OrderedDict
import yfinance as yf
import requests


class ReadXLSException(Exception): pass

class DataFetch(object):

	DEFAULT_HOST_DATA_FOLDERPATH =  "C:\\Users\\%s\\project_data\\finance_web" % getpass.getuser()
	DEFAULT_DATA_FOLDERPATH =  os.getenv("DATA_PATH", DEFAULT_HOST_DATA_FOLDERPATH)
	# DEFAULT_SOURCE_FILENAME = "加權指數歷史資料2000-2025.xlsx"
	DEFAULT_TIME_FIELD_NAME = "時間"	
	DEFAULT_CLOSING_PRICE_FIELD_NAME = "收盤價"	
	# DEFAULT_CONFIG_FOLDERPATH =  "C:\\Users\\%s" % os.getlogin()

	DEFAULT_DATA_DATE_TITLE = "時間"
	DEFAULT_DATA_OPEN_TITLE = "開盤價"
	DEFAULT_DATA_HIGH_TITLE = "最高價"
	DEFAULT_DATA_LOW_TITLE = "最低價"
	DEFAULT_DATA_CLOSE_TITLE = "收盤價"
	DEFAULT_DATA_VOLUME_TITLE = "成交量"
	DEFAULT_DATA_TITLE_LIST = [
		DEFAULT_DATA_DATE_TITLE,
		DEFAULT_DATA_OPEN_TITLE,
		DEFAULT_DATA_HIGH_TITLE,
		DEFAULT_DATA_LOW_TITLE,
		DEFAULT_DATA_CLOSE_TITLE,
		DEFAULT_DATA_VOLUME_TITLE
	]
	DEFAULT_DATA_DATE_TITLE_INDEX = DEFAULT_DATA_TITLE_LIST.index(DEFAULT_DATA_DATE_TITLE)
	DEFAULT_DATA_OPEN_TITLE_INDEX = DEFAULT_DATA_TITLE_LIST.index(DEFAULT_DATA_OPEN_TITLE)
	DEFAULT_DATA_HIGH_TITLE_INDEX = DEFAULT_DATA_TITLE_LIST.index(DEFAULT_DATA_HIGH_TITLE)
	DEFAULT_DATA_LOW_TITLE_INDEX = DEFAULT_DATA_TITLE_LIST.index(DEFAULT_DATA_LOW_TITLE)
	DEFAULT_DATA_CLOSE_TITLE_INDEX = DEFAULT_DATA_TITLE_LIST.index(DEFAULT_DATA_CLOSE_TITLE)
	DEFAULT_DATA_VOLUME_TITLE_INDEX = DEFAULT_DATA_TITLE_LIST.index(DEFAULT_DATA_VOLUME_TITLE)

# Yahoo
	DEFAULT_YAHOO_DATE_TITLE = "Date"
	DEFAULT_YAHOO_OPEN_TITLE = "Open"
	DEFAULT_YAHOO_HIGH_TITLE = "High"
	DEFAULT_YAHOO_LOW_TITLE = "Low"
	DEFAULT_YAHOO_CLOSE_TITLE = "Close"
	DEFAULT_YAHOO_VOLUME_TITLE = "Volume"
	DEFAULT_YAHOO_DATA_TITLE_MAPPING = [
		(DEFAULT_YAHOO_DATE_TITLE, DEFAULT_DATA_DATE_TITLE), 
		(DEFAULT_YAHOO_OPEN_TITLE, DEFAULT_DATA_OPEN_TITLE), 
		(DEFAULT_YAHOO_HIGH_TITLE, DEFAULT_DATA_HIGH_TITLE), 
		(DEFAULT_YAHOO_LOW_TITLE, DEFAULT_DATA_LOW_TITLE), 
		(DEFAULT_YAHOO_CLOSE_TITLE, DEFAULT_DATA_CLOSE_TITLE), 
		(DEFAULT_YAHOO_VOLUME_TITLE, DEFAULT_DATA_VOLUME_TITLE)
	]
	DEFAULT_YAHOO_TITLE_LIST = [x[0] for x in DEFAULT_YAHOO_DATA_TITLE_MAPPING]
	DEFAULT_YAHOO_TITLE_DESCRIPTIOIN_LIST = [x[1] for x in DEFAULT_YAHOO_DATA_TITLE_MAPPING]
	DEFAULT_YAHOO_DATE_TITLE_INDEX = DEFAULT_YAHOO_TITLE_LIST.index(DEFAULT_YAHOO_DATE_TITLE)
	DEFAULT_YAHOO_OPEN_TITLE_INDEX = DEFAULT_YAHOO_TITLE_LIST.index(DEFAULT_YAHOO_OPEN_TITLE)
	DEFAULT_YAHOO_HIGH_TITLE_INDEX = DEFAULT_YAHOO_TITLE_LIST.index(DEFAULT_YAHOO_HIGH_TITLE)
	DEFAULT_YAHOO_LOW_TITLE_INDEX = DEFAULT_YAHOO_TITLE_LIST.index(DEFAULT_YAHOO_LOW_TITLE)
	DEFAULT_YAHOO_CLOSE_TITLE_INDEX = DEFAULT_YAHOO_TITLE_LIST.index(DEFAULT_YAHOO_CLOSE_TITLE)
	DEFAULT_YAHOO_VOLUME_TITLE_INDEX = DEFAULT_YAHOO_TITLE_LIST.index(DEFAULT_YAHOO_VOLUME_TITLE)
	DEFAULT_YAHOO_DATE_FORMAT = "%Y-%m-%d"
# Fin Mind
	DEFAULT_FINMIND_DATE_TITLE = "date"
	DEFAULT_FINMIND_OPEN_TITLE = "open"
	DEFAULT_FINMIND_HIGH_TITLE = "max"
	DEFAULT_FINMIND_LOW_TITLE = "min"
	DEFAULT_FINMIND_CLOSE_TITLE = "close"
	DEFAULT_FINMIND_VOLUME_TITLE = "Trading_Volume"
	DEFAULT_FINMIND_DATA_TITLE_MAPPING = [
		(DEFAULT_FINMIND_DATE_TITLE, DEFAULT_DATA_DATE_TITLE), 
		(DEFAULT_FINMIND_OPEN_TITLE, DEFAULT_DATA_OPEN_TITLE), 
		(DEFAULT_FINMIND_HIGH_TITLE, DEFAULT_DATA_HIGH_TITLE), 
		(DEFAULT_FINMIND_LOW_TITLE, DEFAULT_DATA_LOW_TITLE), 
		(DEFAULT_FINMIND_CLOSE_TITLE, DEFAULT_DATA_CLOSE_TITLE), 
		(DEFAULT_FINMIND_VOLUME_TITLE, DEFAULT_DATA_VOLUME_TITLE)
	]
	DEFAULT_FINMIND_TITLE_LIST = [x[0] for x in DEFAULT_FINMIND_DATA_TITLE_MAPPING]
	DEFAULT_FINMIND_TITLE_DESCRIPTIOIN_LIST = [x[1] for x in DEFAULT_FINMIND_DATA_TITLE_MAPPING]
	DEFAULT_FINMIND_DATE_TITLE_INDEX = DEFAULT_FINMIND_TITLE_LIST.index(DEFAULT_FINMIND_DATE_TITLE)
	DEFAULT_FINMIND_OPEN_TITLE_INDEX = DEFAULT_FINMIND_TITLE_LIST.index(DEFAULT_FINMIND_OPEN_TITLE)
	DEFAULT_FINMIND_HIGH_TITLE_INDEX = DEFAULT_FINMIND_TITLE_LIST.index(DEFAULT_FINMIND_HIGH_TITLE)
	DEFAULT_FINMIND_LOW_TITLE_INDEX = DEFAULT_FINMIND_TITLE_LIST.index(DEFAULT_FINMIND_LOW_TITLE)
	DEFAULT_FINMIND_CLOSE_TITLE_INDEX = DEFAULT_FINMIND_TITLE_LIST.index(DEFAULT_FINMIND_CLOSE_TITLE)
	DEFAULT_FINMIND_VOLUME_TITLE_INDEX = DEFAULT_FINMIND_TITLE_LIST.index(DEFAULT_FINMIND_VOLUME_TITLE)
	DEFAULT_FINMIND_DATE_FORMAT = "%Y-%m-%d"
	DEFAULT_FINMIND_MIN_DATE = date(1900, 1, 1)
	DEFAULT_FINMIND_MIN_DATE_STR = DEFAULT_FINMIND_MIN_DATE.strftime(DEFAULT_FINMIND_DATE_FORMAT)

	DEFAULT_DATE_BASE_NUMBER = 36526
	DEFAULT_DATE_BASE = date(2000, 1, 1)
	# DEFAULT_MIN_DATE = date(1900, 1, 1)
	# DEFAULT_MAX_DATE = date(2099, 12, 31)
	DEFAULT_YAHOO_TODAY_DATA_UPDATE_HOUR = 15
	DEFAULT_WARNING_MEWSAGE_PREFIX = "WARNING"

	@classmethod
	def __is_string(cls, value):
		is_string = False
		try:
			int(value)
		except ValueError:
			is_string = True
		return is_string


	@classmethod
	def __check_file_exist(cls, filepath):
		check_exist = True
		try:
			os.stat(filepath)
		except OSError as exception:
			if exception.errno != errno.ENOENT:
				print ("%s: %s" % (errno.errorcode[exception.errno], os.strerror(exception.errno)))
				raise
			check_exist = False
		return check_exist


	@classmethod
	def __check_date_in_range(cls, cur_date, date_range_start, date_range_end):
		if isinstance(cur_date, str):
			cur_date = cls.__date_str2obj(cur_date)
		if isinstance(date_range_start, str):
			date_range_start = cls.__date_str2obj(date_range_start)
		if isinstance(date_range_end, str):
			date_range_end = cls.__date_str2obj(date_range_end)
		if date_range_start <= cur_date <= date_range_end:
			return True
		return False


# ========= XLSX 工具 =========
	@classmethod
	def __read_xlsx(cls, source_filepath):
		# wb = self.__get_workbook()
		wb = load_workbook(source_filepath)
		ws = wb.active
		rows = []
		header = [c.value for c in ws[1]]
		rows.append(header)
		# for row in ws.iter_rows(min_row=2, values_only=True):
		# 	rows.append(dict(zip(header, row)))
		for row in ws.iter_rows(min_row=2, values_only=True):
			rows.append(row)
		return rows


	@classmethod
	def __write_xlsx(cls, source_filepath, rows, refresh_data=False):
		ws = None
		start_index = None
		# import pdb; pdb.set_trace()
		if not refresh_data and cls.__check_file_exist(source_filepath):
# If file exists, append new data...
			wb = load_workbook(source_filepath)
			ws = wb.active
# openpyxl 的 row / column 是從 1 開始，不是 0
# openpyxl 是 Excel 操作庫，不是資料結構庫 它選擇「跟 Excel 一樣」而不是「跟 Python 一樣」
# Excel 的世界本來就是從 1 開始
			data_last_date_str = ws.cell(row=ws.max_row, column=cls.DEFAULT_YAHOO_DATE_TITLE_INDEX + 1).value
			data_last_date = datetime.strptime(data_last_date_str, cls.DEFAULT_YAHOO_DATE_FORMAT).date()
			for index, row in enumerate(rows[1:]):
				row_date_str = row[cls.DEFAULT_YAHOO_DATE_TITLE_INDEX]
				row_date = datetime.strptime(row_date_str, cls.DEFAULT_YAHOO_DATE_FORMAT).date()
				if row_date > data_last_date:
# row 0 是 header，所以 index 從 0 開始對應到 rows[1:] 的第一行資料
					start_index = index + 1 # 因為 rows[1:] 的 index 是從 0 開始，所以要加 1 才是正確的行數
					break
			# if start_index is None: # 已存在，不寫入
			# 	wb.close()
			# 	return
		else:
# If file does NOT exist, create new file...
			wb = Workbook()
			ws = wb.active
			# headers = rows[0].keys()
			# ws.append(list(headers))
			ws.append(cls.DEFAULT_DATA_TITLE_LIST)
			start_index = 1
		# import pdb; pdb.set_trace()
		if start_index != None:
			for r in rows[start_index:]:
				ws.append(r)
			data_len = len(rows[start_index:])
			print(f"{data_len} is updated in {source_filepath}")
			wb.save(source_filepath)
		else:
			print(f"No data is updated in {source_filepath}")
		wb.close()


	@classmethod
	def __is_excel_locked(cls, filepath):
# Excel lock file
		lock_path = os.path.join(os.path.dirname(filepath), "~$" + os.path.basename(filepath))
		if os.path.exists(lock_path):
			return True
# OS lock
		try:
			with open(filepath, "a"):
				return False
		except PermissionError:
			return True


	@classmethod
	def __get_yahoo_raw_data(cls, stock_symbol, fetch_start=None, fetch_end=None):
# 抓取歷史資料
# start=None: 會自動設為一個很早的日期（實務上接近 1900-01-01）等同於「從資料能取得的最早時間開始抓」
# end=None: 會自動設為「現在時間」（today / now）
# start=None 且 end=None: 抓「該股票所有可用歷史資料」 
		# hist = yf.download(stock_symbol, start=fetch_start.strftime(self.DEFAULT_YAHOO_DATE_FORMAT), end=fetch_end.strftime(self.DEFAULT_YAHOO_DATE_FORMAT), group_by='column')
		# hist = yf.download(stock_symbol, start=fetch_start, end=fetch_end, group_by='column')
		if fetch_start is None and fetch_end is None:
# 全部資料
			hist = yf.download(stock_symbol, period="max", group_by='column')
		elif fetch_start is None and fetch_end is not None:
# 從最早到 end
			hist = yf.download(stock_symbol, period="max", end=fetch_end, group_by='column')
		elif fetch_start is not None and fetch_end is None:
# 從 start 到最新
			hist = yf.download(stock_symbol, start=fetch_start, group_by='column')
		else:
			hist = yf.download(stock_symbol, start=fetch_start, end=fetch_end, group_by='column')
		if hist.empty:
			return hist
		# import pdb; pdb.set_trace()
		hist = hist.rename(columns={
			cls.DEFAULT_YAHOO_DATE_TITLE: cls.DEFAULT_DATA_DATE_TITLE,
			cls.DEFAULT_YAHOO_OPEN_TITLE: cls.DEFAULT_DATA_OPEN_TITLE,
			cls.DEFAULT_YAHOO_HIGH_TITLE: cls.DEFAULT_DATA_HIGH_TITLE,
			cls.DEFAULT_YAHOO_LOW_TITLE: cls.DEFAULT_DATA_LOW_TITLE,
			cls.DEFAULT_YAHOO_CLOSE_TITLE: cls.DEFAULT_DATA_CLOSE_TITLE,
			cls.DEFAULT_YAHOO_VOLUME_TITLE: cls.DEFAULT_DATA_VOLUME_TITLE,
		})
		hist = hist.reset_index()
# 把 MultiIndex 欄位轉成單層欄位
		if isinstance(hist.columns, pd.MultiIndex):
			hist.columns = hist.columns.get_level_values(0)
		return hist[[cls.DEFAULT_DATA_DATE_TITLE, cls.DEFAULT_DATA_OPEN_TITLE, cls.DEFAULT_DATA_HIGH_TITLE, cls.DEFAULT_DATA_LOW_TITLE, cls.DEFAULT_DATA_CLOSE_TITLE, cls.DEFAULT_DATA_VOLUME_TITLE]]
		# return list[cls.DEFAULT_DATA_TITLE_LIST]


	@classmethod
	def __get_yahoo_data(cls, stock_symbol, fetch_start=None, fetch_end=None):
		hist = cls.__get_yahoo_raw_data(stock_symbol, fetch_start, fetch_end)
# 轉成 CSV 格式
		row_data_list = []
		row_data_list.append(cls.DEFAULT_DATA_TITLE_LIST)
		# import pdb; pdb.set_trace()
		for _, r in hist.iterrows():
			one_row_data = []
			for data_title in cls.DEFAULT_DATA_TITLE_LIST:
				if data_title == cls.DEFAULT_DATA_DATE_TITLE:
					date_str = r[data_title].strftime(cls.DEFAULT_YAHOO_DATE_FORMAT)
					one_row_data.append(date_str)
				else:
					one_row_data.append(r[data_title])
# Check if fake row
			if one_row_data[cls.DEFAULT_YAHOO_VOLUME_TITLE_INDEX] == 0:
				[o, h, l, c] = [one_row_data[i] for i in [cls.DEFAULT_YAHOO_OPEN_TITLE_INDEX, cls.DEFAULT_YAHOO_HIGH_TITLE_INDEX, cls.DEFAULT_YAHOO_LOW_TITLE_INDEX, cls.DEFAULT_YAHOO_CLOSE_TITLE_INDEX]]
				if not any(math.isnan(x) for x in [o,h,l,c]):
					if o == h == l == c:
						continue
			row_data_list.append(one_row_data)
		# import pdb; pdb.set_trace()
		latest_data_index = None
		if len(row_data_list) > 0:
			latest_date_str = row_data_list[-1][cls.DEFAULT_YAHOO_DATE_TITLE_INDEX]
			latest_date = datetime.strptime(latest_date_str, cls.DEFAULT_YAHOO_DATE_FORMAT).date()
			today = datetime.today().date()
# skip today's data
			if latest_date == today:
# 				if datetime.now() < datetime(today.year, today.month, today.day, cls.DEFAULT_YAHOO_TODAY_DATA_UPDATE_HOUR):
# # Today's data is not updated yet, so skip today's data
# 					latest_data_index = -1
				latest_data_index = -1
		if latest_data_index is not None:
			row_data_list = row_data_list[:latest_data_index]
		return row_data_list


	@classmethod
	def __get_finmind_raw_data(cls, stock_symbol, fetch_start=None, fetch_end=None, token=None):
		url = "https://api.finmindtrade.com/api/v4/data"
		params = {
			"dataset": "TaiwanStockPrice",
			"data_id": stock_symbol,
		}
		if fetch_start is not None:
			params["start_date"] = fetch_start.strftime(cls.DEFAULT_FINMIND_DATE_FORMAT)
		else:
			params["start_date"] = cls.DEFAULT_FINMIND_MIN_DATE_STR
		if fetch_end is not None:
			params["end_date"] = fetch_end.strftime(cls.DEFAULT_FINMIND_DATE_FORMAT)
		if token is not None:
			params["token"] = token
		# import pdb; pdb.set_trace()
		resp = requests.get(url, params=params).json()
		# print(dir(resp))
		if resp["status"] != 200:
			print("ERROR: Fails to fetch data[%s], due to: %s" % (stock_symbol, resp["msg"]))
			return pd.DataFrame([])
		hist = pd.DataFrame(resp["data"])
		# if hist.empty:
		# 	return hist
		hist = hist.rename(columns={
			cls.DEFAULT_FINMIND_DATE_TITLE: cls.DEFAULT_DATA_DATE_TITLE,
			cls.DEFAULT_FINMIND_OPEN_TITLE: cls.DEFAULT_DATA_OPEN_TITLE,
			cls.DEFAULT_FINMIND_HIGH_TITLE: cls.DEFAULT_DATA_HIGH_TITLE,
			cls.DEFAULT_FINMIND_LOW_TITLE: cls.DEFAULT_DATA_LOW_TITLE,
			cls.DEFAULT_FINMIND_CLOSE_TITLE: cls.DEFAULT_DATA_CLOSE_TITLE,
			cls.DEFAULT_FINMIND_VOLUME_TITLE: cls.DEFAULT_DATA_VOLUME_TITLE,
		})
		hist[cls.DEFAULT_DATA_DATE_TITLE] = pd.to_datetime(hist[cls.DEFAULT_DATA_DATE_TITLE])
		hist = hist.set_index(cls.DEFAULT_DATA_DATE_TITLE).sort_index()
		return hist[[cls.DEFAULT_DATA_OPEN_TITLE, cls.DEFAULT_DATA_HIGH_TITLE, cls.DEFAULT_DATA_LOW_TITLE, cls.DEFAULT_DATA_CLOSE_TITLE, cls.DEFAULT_DATA_VOLUME_TITLE]]


	@classmethod
	def __get_finmind_data(cls, stock_symbol, fetch_start=None, fetch_end=None, token=None):
		hist = cls.__get_finmind_raw_data(stock_symbol, fetch_start, fetch_end, token)
# 轉成 CSV 格式
		row_data_list = []
		row_data_list.append(cls.DEFAULT_DATA_TITLE_LIST)
		# import pdb; pdb.set_trace()
		for idx, row in hist.iterrows():
			one_row_data = []
			for data_title in cls.DEFAULT_DATA_TITLE_LIST:
				if data_title == cls.DEFAULT_DATA_DATE_TITLE:
					date_str = idx.strftime(cls.DEFAULT_FINMIND_DATE_FORMAT)
					one_row_data.append(date_str)
				else:
					one_row_data.append(row[data_title])
# # Check if fake row
# 			if one_row_data[cls.DEFAULT_FINMIND_VOLUME_TITLE_INDEX] == 0:
# 				[o, h, l, c] = [one_row_data[i] for i in [cls.DEFAULT_FINMIND_OPEN_TITLE_INDEX, cls.DEFAULT_FINMIND_HIGH_TITLE_INDEX, cls.DEFAULT_FINMIND_LOW_TITLE_INDEX, cls.DEFAULT_FINMIND_CLOSE_TITLE_INDEX]]
# 				if not any(math.isnan(x) for x in [o,h,l,c]):
# 					if o == h == l == c:
# 						continue
			row_data_list.append(one_row_data)
		return row_data_list


	def __init__(self, cfg):
		self.xcfg = {
			"source_folderpath": None,
			# "source_filename": None,
			"stock_symbol_string": None,
			"data_date_range_string": None,
			"refresh_data": False,
			# "date_range_start": None,
			# "date_range_end": None,
			"show_warning": False,
			"finmind_token": None,
			"fetch_method_string": None,
		}
		# import pdb; pdb.set_trace()
		self.xcfg.update(cfg)
		self.xcfg["source_folderpath"] = self.DEFAULT_DATA_FOLDERPATH if self.xcfg["source_folderpath"] is None else self.xcfg["source_folderpath"]
		self.workbook = None

		self.filepath_dict = OrderedDict()
		self.filepath_dict["source_folderpath"] = self.xcfg["source_folderpath"]
# Fetch Method: 
# 0: Auto Select. 1: FinMind. 2: Yahoo
		if self.xcfg["fetch_method_string"] is None:
			self.fetch_method = 0
		else:
			mobj = re.match(('[\d]'), self.xcfg["fetch_method_string"])
			if mobj is not None:
				fetch_method = int(mobj.group(0))
				if fetch_method not in [0, 1, 2,]:
					raise ValueError("Unknown fetch method: %d" % fetch_method)
				self.fetch_method = fetch_method
			else:	
				if re.match("default", self.xcfg["fetch_method_string"], re.I):
					self.fetch_method = 0
				elif re.match("finmind", self.xcfg["fetch_method_string"], re.I):
					self.fetch_method = 1
				elif re.match("yahoo", self.xcfg["fetch_method_string"], re.I):
					self.fetch_method = 2
				else:
					raise ValueError("Unknown fetch method: %s" % self.xcfg["fetch_method_string"])


	def __enter__(self):
		return self


	def __exit__(self, type, msg, traceback):
		return False


	def __date_str2list(self, date_str, skip_year=False):
		# import pdb; pdb.set_trace()
		# print(date_str)
		if date_str.find("/") != -1:
			elem_list = date_str.split("/")
			if len(elem_list) == 2:
				[year, month, day] = [self.cur_year, int(elem_list[0]), int(elem_list[1])]
			elif len(elem_list) == 3:
				[month, day, year] = list(map(int, elem_list))
			else:
				raise ValueError("Incorrect date string format: %s" % date_str)
		elif date_str.find("-") != -1:
			elem_list = date_str.split("-")
			if len(elem_list) == 2:
				[year, month, day] = [self.cur_year, int(elem_list[0]), int(elem_list[1])]
			elif len(elem_list) == 3:
				[year, month, day] = list(map(int, elem_list))
			else:
				raise ValueError("Incorrect date string format: %s" % date_str)
		else:
			raise ValueError("Incorrect date string format: %s" % date_str)
		return [month, day] if skip_year else [year, month, day]


	def __date_str2obj(self, date_str):
		[year, month, day] = self.__date_str2list(date_str)
		date_obj = None
		try:
			date_obj = date(year, month, day)
		except Exception as e:
			print("Unsupport date format[%s] due to: %s" % (date_str, str(e)))
			# import pdb; pdb.set_trace()
			raise e
		return date_obj


	def __date_number2obj(self, date_number):
		day_diff = int(date_number) - self.DEFAULT_DATE_BASE_NUMBER
		date_obj = self.DEFAULT_DATE_BASE + timedelta(days=day_diff)
		return date_obj


	def __date_number2list(self, date_number):
		date_obj = self.__date_number2obj(date_number)
		return [date_obj.year, date_obj.month, date_obj.day]


	def __get_data(self, stock_symbol, get_start=None, get_end=None):
# 0: Auto Select. 1: FinMind. 2: Yahoo.
		use_finmind = False
		if self.fetch_method == 0:
			if stock_symbol.endswith(".TW"):
				use_finmind = True
		elif self.fetch_method == 1:
			use_finmind = True
		hist_data = None
		if use_finmind:
			if self.xcfg["finmind_token"] is None:
				raise ValueError("Fin Mind Token should NOT be None")
			hist_data = self.__get_finmind_data(stock_symbol.rstrip(".TW"), get_start, get_end, self.xcfg["finmind_token"])
		else:
			hist_data = self.__get_yahoo_data(stock_symbol, get_start, get_end)
		return hist_data


	def __fetch_data(self, stock_symbol, date_range_start_str=None, date_range_end_str=None):
		"""
		取得歷史資料，如果本地已存在 CSV，則只抓最新日期後的資料。
		"""
		# csv_file = os.path.join(self.xcfg["source_filepath"], f"{stock_symbol}.csv")
# 檢查是否已存在本地資料
		# import pdb; pdb.set_trace()
		source_filepath = os.path.join(self.xcfg["source_folderpath"], f"{stock_symbol}.xlsx")
		file_exist = self.__check_file_exist(source_filepath)
		if file_exist:
			if self.__is_excel_locked(source_filepath):
				return f"ERROR: The file {source_filepath} is locked by other process, so skip fetching data for {stock_symbol}..."
		fetch_start = fetch_end = None
		refresh_data = self.xcfg["refresh_data"]
		# import pdb; pdb.set_trace()
		if file_exist:
			date_range_start = datetime.strptime(date_range_start_str, self.DEFAULT_YAHOO_DATE_FORMAT).date() if date_range_start_str is not None else None  # self.DEFAULT_MIN_DATE  # first_date
			date_range_end = datetime.strptime(date_range_end_str, self.DEFAULT_YAHOO_DATE_FORMAT).date() if date_range_end_str is not None else None  # self.DEFAULT_MAX_DATE  # datetime.today().date()  # last_date
			# import pdb; pdb.set_trace()
			if refresh_data:
				fetch_start = date_range_start
			else:	
				rows = self.__read_xlsx(source_filepath)
				first_date_str = rows[1][self.DEFAULT_YAHOO_DATE_TITLE_INDEX]
				first_date = datetime.strptime(first_date_str, self.DEFAULT_YAHOO_DATE_FORMAT).date()
				last_date_str = rows[-1][self.DEFAULT_YAHOO_DATE_TITLE_INDEX]
				last_date = datetime.strptime(last_date_str, self.DEFAULT_YAHOO_DATE_FORMAT).date()
# Check the boundary condition of date range
				if (date_range_start is not None) and (date_range_end is not None) and (date_range_start > date_range_end):
					return f"ERROR: The start date {date_range_start_str} is later than the end date {date_range_end_str}."
				if (date_range_end is not None) and (date_range_end < (first_date - timedelta(days=1))):
					return f"ERROR: The end date {date_range_end_str} is out of boundary of the local data {first_date_str} - {last_date_str}."
				if (date_range_start is not None) and (date_range_start > (last_date + timedelta(days=1))):
					return f"ERROR: The start date {date_range_start_str} is out of boundary of the local data {first_date_str} - {last_date_str}."
				if date_range_start is not None:
					if date_range_start < first_date:
						refresh_data = True
					elif (date_range_end is not None) and first_date <= date_range_end <= last_date:
						return f"WARNING: The date range {date_range_start} - {date_range_end} is within the local data."
				fetch_start = last_date + timedelta(days=1)
		else:
			if date_range_start_str is not None:
				fetch_start = datetime.strptime(date_range_start_str, self.DEFAULT_YAHOO_DATE_FORMAT).date()
		if date_range_end_str is not None:
			fetch_end = datetime.strptime(date_range_end_str, self.DEFAULT_YAHOO_DATE_FORMAT).date()
			if fetch_end > datetime.today().date():
				return f"ERROR: The end date {date_range_end_str} should NOT be later than today"
		if (fetch_start is not None) and (fetch_end is not None):
			if fetch_start > fetch_end:
				return f"ERROR: Incorrect time range %s - %s" % (fetch_start.strftime(self.DEFAULT_YAHOO_DATE_FORMAT), fetch_end.strftime(self.DEFAULT_YAHOO_DATE_FORMAT))
# 如果已經最新，直接返回
		# import pdb; pdb.set_trace()
		if not refresh_data and fetch_start is not None and fetch_end is None:
			no_latest_data = False
			today = datetime.today().date()
			if fetch_start > today:
				no_latest_data = True
			elif fetch_start == today and datetime.now() < datetime(today.year, today.month, today.day, self.DEFAULT_YAHOO_TODAY_DATA_UPDATE_HOUR):
				no_latest_data = True
			if no_latest_data:
				return f"WARNING: The data of {stock_symbol} is already the latest."
# 抓取歷史資料
		# import pdb; pdb.set_trace()
		row_data_list = self.__get_data(stock_symbol, fetch_start, fetch_end)
		if len(row_data_list) == 0:
			return f"WARNING: No new data in {stock_symbol}."
		self.__write_xlsx(source_filepath, row_data_list, refresh_data)
		return None


	def fetch_data(self):
		# import pdb; pdb.set_trace()
		if self.xcfg["stock_symbol_string"] is None:
			print("Warning: No stock to fetch...")
			return
		stock_symbol_list = self.xcfg["stock_symbol_string"].split(",")
		date_range_start_str = date_range_end_str = None
		if self.xcfg["data_date_range_string"] is not None:
			date_range_elems = self.xcfg["data_date_range_string"].split(":")
			if len(date_range_elems) == 2:
				if date_range_elems[0] != "":
					date_range_start_str = date_range_elems[0]
				if date_range_elems[1] != "":
					date_range_end_str = date_range_elems[1]
			else:
				print("Error: Incorrect date range format[%s]" % self.xcfg["data_date_range_string"])
				return 
		for stock_symbol in stock_symbol_list:
			return_message = self.__fetch_data(stock_symbol, date_range_start_str, date_range_end_str)
			if return_message is not None:
				if not self.xcfg["show_warning"]:
					if return_message.startswith(self.DEFAULT_WARNING_MEWSAGE_PREFIX):
						continue
				print(f"No {stock_symbol} data are fetched, due to: {return_message}")


	def __inspect_data(self, stock_symbol):
		# import pdb; pdb.set_trace()
		source_filepath = os.path.join(self.xcfg["source_folderpath"], f"{stock_symbol}.xlsx")
		file_exist = self.__check_file_exist(source_filepath)
		data_info = None
		if file_exist:
			if self.__is_excel_locked(source_filepath):
				# print(f"ERROR: The file {source_filepath} is locked by other process, fails to show data info for {stock_symbol}...")
				raise ReadXLSException(f"The file {source_filepath} is locked by other process, fails to inspect data info for {stock_symbol}...")
			rows = self.__read_xlsx(source_filepath)
			data_info = {
				"first_date_str": rows[1][self.DEFAULT_YAHOO_DATE_TITLE_INDEX], 
				"last_date_str": rows[-1][self.DEFAULT_YAHOO_DATE_TITLE_INDEX], 
				"data_cnt": len(rows) - 1,
			}
		return data_info


	def show_data_info(self):
		# import pdb; pdb.set_trace()
		if self.xcfg["stock_symbol_string"] is None:
			print("Warning: No stock to fetch...")
			return
		stock_symbol_list = self.xcfg["stock_symbol_string"].split(",")
		for stock_symbol in stock_symbol_list:
			try:
				data_info = self.__inspect_data(stock_symbol)
				if data_info is not None:
					first_date_str = data_info["first_date_str"]
					last_date_str = data_info["last_date_str"]
					data_cnt = data_info["data_cnt"]
					print(f"{stock_symbol}: {data_cnt} data from {first_date_str} to {last_date_str}")
				else:
					print(f"{stock_symbol}: No data exists...")
			except ReadXLSException as e:
				print(str(e))


	def print_filepath(self):
		print("************** File Path **************")
		for key, value in self.filepath_dict.items():
			print("%s: %s" % (key, value))


if __name__ == "__main__":
# argparse 預設會把 help 文字裡的換行與多重空白「壓縮」成一行，所以你在字串裡寫的 \n 不一定會照原樣顯示。 => 建立 parser 時加上 formatter_class=argparse.RawTextHelpFormatter
	parser = argparse.ArgumentParser(description='Print help', formatter_class=argparse.RawTextHelpFormatter)
	'''
	參數基本上分兩種，一種是位置參數 (positional argument)，另一種就是選擇性參數 (optional argument)
	* example2.py
	parser.add_argument("pos1", help="positional argument 1")
	parser.add_argument("-o", "--optional-arg", help="optional argument", dest="opt", default="default")

	# python example2.py hello -o world 
	positional arg: hello
	optional arg: world
	'''
# How to add option without any argument? use action='store_true'
	'''
	'store_true' and 'store_false' - ?些是 'store_const' 分?用作存? True 和 False 值的特殊用例。
	另外，它?的默?值分?? False 和 True。例如:

	>>> parser = argparse.ArgumentParser()
	>>> parser.add_argument('--foo', action='store_true')
	>>> parser.add_argument('--bar', action='store_false')
	>>> parser.add_argument('--baz', action='store_false')
	'''
	parser.add_argument('--source_folderpath', required=False, help='Fetch data into the XLS files in the designated folder path. Ex: %s' % DataFetch.DEFAULT_DATA_FOLDERPATH)
	parser.add_argument('-f', '--fetch_data', required=False, action='store_true', help='Fetch the data of the specific target and exit.')
	parser.add_argument('--stock_symbol_list', required=False, help='The stock symbol list. Mutiple stock symbols are split by comma. Ex: 00850.TW,00881.TW,00692.TW,MSFT,GOOG')
	parser.add_argument('--data_date_range', required=False, 
		 help='''The data during the date range.
  Date range:
    Format: yy1-mm1-dd1:yy2-mm2-dd2   From yy1-mm1-dd1 to yy2-mm2-dd2   Ex: 2014-09-04:2025-10-15
    Format: yy-mm-dd:   From yy-mm-dd to 'the last date of the data'   Ex: 2014-09-04:
    Format: :yy-mm-dd   From 'the first date of the data' to yy-mm-dd   Ex: :2025-09-04
    * Caution: Only take effect when --fetch_data is set.''')
	parser.add_argument('--refresh_data', required=False, action='store_true', help='Ignore the existing the XLSX file so that the data will be fetched from the earliest date to today. Only take effect when --fetch_data is set.')
	parser.add_argument('--show_data_info', required=False, action='store_true', help='Show data info and exit.')
	parser.add_argument('--print_filepath', required=False, action='store_true', help='Print the filepaths used in the process and exit.')
	parser.add_argument('--show_warning', required=False, action='store_true', help='Show warnings messages.')
	parser.add_argument('--finmind_token', required=False, help='The FinMind Token')
	parser.add_argument('--fetch_method', required=False, help='Select the fetch method: Auto Select[0]/FinMind[1]/Yahoo[2]. Default: auto select. Use FinMind to fetch as stock symbol is .TW as suffix, otherwise use Yahoo to fetch.')
	args = parser.parse_args()
	# import pdb; pdb.set_trace()
	cfg = {}
	if args.source_folderpath is not None: cfg['source_folderpath'] = args.source_folderpath
	if args.stock_symbol_list is not None: cfg['stock_symbol_string'] = args.stock_symbol_list
	if args.data_date_range is not None: cfg['data_date_range_string'] = args.data_date_range
	if args.refresh_data: cfg['refresh_data'] = True
	if args.show_warning: cfg['show_warning'] = True
	if args.finmind_token is not None: cfg['finmind_token'] = args.finmind_token
	if args.fetch_method is not None: cfg['fetch_method_string'] = args.fetch_method
	# import pdb; pdb.set_trace()
	with DataFetch(cfg) as obj:
		if args.show_data_info:
			obj.show_data_info()
			sys.exit(0)
		if args.print_filepath:
			obj.print_filepath()
			sys.exit(0)
		if args.fetch_data:
			obj.fetch_data()
			sys.exit(0)