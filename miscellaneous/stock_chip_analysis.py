#! /usr/bin/python
# -*- coding: utf8 -*-

import os
import sys
import re
import errno
# # '''
# # Question: How to Solve xlrd.biffh.XLRDError: Excel xlsx file; not supported ?
# # Answer : The latest version of xlrd(2.01) only supports .xls files. Installing the older version 1.2.0 to open .xlsx files.
# # '''
# import xlrd
# # import xlsxwriter
from openpyxl import Workbook, load_workbook
import argparse
import copy
import csv
import json
from datetime import datetime
import getpass
# from pymongo import MongoClient
from collections import OrderedDict


class StockChipAnalysis(object):

	DEFAULT_HOST_ROOT_FOLDERPATH =  "C:\\Users\\%s\\project_data\\finance_web" % getpass.getuser()
	# DEFAULT_SOURCE_FOLDERPATH =  "C:\\Users\\%s\\Downloads" % os.getlogin()
	DEFAULT_HOST_SOURCE_FOLDERPATH =  DEFAULT_HOST_ROOT_FOLDERPATH  # "C:\\Users\\%s\\Downloads" % getpass.getuser()
	DEFAULT_SOURCE_FOLDERPATH =  os.getenv("DATA_PATH", DEFAULT_HOST_SOURCE_FOLDERPATH)
	DEFAULT_SOURCE_FILENAME = "stock_chip_analysis"
	DEFAULT_SOURCE_FULL_FILENAME = "%s.xlsm" % DEFAULT_SOURCE_FILENAME
	# DEFAULT_CONFIG_FOLDERPATH =  "C:\\Users\\%s" % os.getlogin()
	DEFAULT_HOST_CONFIG_FOLDERPATH =  DEFAULT_HOST_ROOT_FOLDERPATH  # "C:\\Users\\%s" % getpass.getuser()
	DEFAULT_CONFIG_FOLDERPATH =  os.getenv("DATA_PATH", DEFAULT_HOST_CONFIG_FOLDERPATH)
	DEFAULT_TRACKED_STOCK_LIST_FILENAME = "chip_analysis_stock_list.txt"
	# DEFAULT_REPORT_FILENAME = "chip_analysis_report.xlsx"
	DEFAULT_OUTPUT_RESULT_FILENAME = "output_result.txt"
	SHEET_METADATA_DICT = {
		# u"SSB": {
		# 	"key_mode": 0, # 2489 瑞軒
		# 	"data_start_column_index": 1,
		# },
		# u"個股夏普值": {
		# 	"key_mode": 0, # 2489 瑞軒
		# 	"data_start_column_index": 1,
		# },
		# u"月平均報酬": {
		# 	"key_mode": 0, # 2489 瑞軒
		# 	"data_start_column_index": 1,
		# },
		u"台股 ETF": {
			"key_mode": 4, # 00727B
			"data_start_column_index": 2,
			"sheet_rows": -1,
			"sheet_columns": 7,
		},
		u"美股 ETF": {
			"key_mode": 5, # JEPQ
			"data_start_column_index": 2,
		},
		# u"外資賺錢": {
		# 	"key_mode": 3, # 台積電(2330)
		# 	"data_start_column_index": 1,
		# },
		u"券商賺錢": {
			"key_mode": 3, # 台積電(2330)
			"data_start_column_index": 1,
		},
		# u"大戶籌碼": {
		# 	"key_mode": 0, # 2489 瑞軒
		# 	"data_start_column_index": 1,
		# },
		# u"成交比重": {
		# 	"key_mode": 0, # 2489 瑞軒
		# 	"data_start_column_index": 1,
		# },
		# u"控盤券商3日買超": {
		# 	"key_mode": 3, # 日月光投控(3711)
		# 	"data_start_column_index": 1,
		# },
		# u"控盤券商3日賣超": {
		# 	"key_mode": 3, # 日月光投控(3711)
		# 	"data_start_column_index": 1,
		# },
		# u"極光波段": {
		# 	"key_mode": 0, # 2489 瑞軒
		# 	"data_start_column_index": 1,
		# },
		# u"短線多空": {
		# 	"key_mode": 0, # 2504 國產
		# 	"data_start_column_index": 1,
		# },
		# u"主法量率": {
		# 	"key_mode": 0, # 2504 國產
		# 	"data_start_column_index": 1,
		# },
		# u"六大買超": {
		# 	"key_mode": 0, # 2504 國產
		# 	"data_start_column_index": 1,
		# },
		# u"大戶持股變化": {
		# 	"key_mode": 0, # 2504 國產
		# 	"data_start_column_index": 1,
		# },
		u"主力買超天數累計": {
			"key_mode": 0, # 2504 國產
			"data_start_column_index": 1,
		},
		# u"法人買超天數累計": {
		# 	"key_mode": 0, # 2504 國產
		# 	"data_start_column_index": 1,
		# },
		u"法人共同買超累計": {
			"key_mode": 1, # 1476
			"data_start_column_index": 2,
			"sheet_columns": 9,
		},
		u"外資買超天數累計": {
			"key_mode": 0, # 2504 國產
			"data_start_column_index": 1,
		},
		u"投信買超天數累計": {
			"key_mode": 0, # 2504 國產
			"data_start_column_index": 1,
		},
		# u"上市融資增加": {
		# 	"key_mode": 2, # 4736  泰博
		# 	"data_start_column_index": 1,
		# },
		# u"上櫃融資增加": {
		# 	"key_mode": 2, # 4736  泰博
		# 	"data_start_column_index": 1,
		# },
	}
	ALL_SHEET_NAME_LIST = SHEET_METADATA_DICT.keys()
	DEFAULT_SHEET_NAME_LIST = [u"台股 ETF", u"美股 ETF", u"券商賺錢", u"主力買超天數累計", u"法人共同買超累計", u"外資買超天數累計", u"投信買超天數累計",]  #  u"個股夏普值", u"月平均報酬", u"六大買超", u"大戶籌碼", u"SSB", u"上市融資增加", u"上櫃融資增加", u"外資賺錢", u"成交比重", u"主法量率",]
	SHEET_SET_LIST = [
		[u"法人共同買超累計", u"主力買超天數累計", u"外資買超天數累計", u"投信買超天數累計",],
		[u"法人共同買超累計", u"外資買超天數累計", u"投信買超天數累計",],
		[u"外資買超天數累計", u"投信買超天數累計",],
	]
	# MONTHLY_AVERAGE_RETURN_SHEETNAME = "月平均報酬"
	DEFAULT_MIN_CONSECUTIVE_OVER_BUY_DAYS = 3
	DEFAULT_MAX_CONSECUTIVE_OVER_BUY_DAYS = 15
	CONSECUTIVE_OVER_BUY_DAYS_SHEETNAME_LIST = [u"主力買超天數累計", u"外資買超天數累計", u"投信買超天數累計",]
	CONSECUTIVE_OVER_BUY_DAYS_FIELDNAME_LIST = [u"累計天數", u"累計天數", u"投信買超累計天數",]
	DEFAULT_MINIMUM_VOLUME = 1000
	MINIMUM_VOLUME_SHEETNAME_LIST = [u"主力買超天數累計", u"外資買超天數累計", u"投信買超天數累計",]
	MINIMUM_VOLUME_FIELDNAME_LIST = [u"主力買超張數", u"外資累計買超張數", u"投信累計買超張數",]
	# DEFAULT_MAIN_FORCE_INSTUITIONAL_INVESTORS_RATIO_THRESHOLD = 10.0
	# DEFAULT_MAIN_FORCE_INSTUITIONAL_INVESTORS_RATIO_CONSECUTIVE_DAYS = 3
	# MAIN_FORCE_INSTUITIONAL_INVESTORS_RATIO_SHEETNAME = "主法量率"
	# MAIN_FORCE_INSTUITIONAL_INVESTORS_RATIO_FIELDNAME = "主力法人佔量率"
	# DEFAULT_STOCK_SHARPE_RATIO_RANKING_PERCENTAGE_THRESHOLD = 20
	# STOCK_SHARPE_RATIO_RANKING_SHEETNAME = "個股夏普值"
	LARGE_SHAREHOLD_POSITION_SHEETNAME = "大戶籌碼"
	LARGE_SHAREHOLD_POSITION_FIELDNAME_SHARPE_RATIO = "夏普值"
	LARGE_SHAREHOLD_POSITION_FIELDNAME_STANDARD_DEVIATION = "標準差"
	SSB_SORT_FIELD_NAME_LIST = ["夏普", "標準差", "貝它",]
	ETF_SHEET_NAME_LIST = ["台股 ETF", "美股 ETF",]
	ETF_SORT_FIELD_NAME_LIST = ["年化標準差", "年報酬", "Alpha", "Beta", "Sharpe", "R-Squared",]
# # CB Related
# 	DEFAULT_CB_FOLDERPATH =  "C:\\可轉債"
# 	DEFAULT_CB_PUBLISH_FILENAME = "可轉債發行"
# 	DEFAULT_CB_PUBLISH_FULL_FILENAME = "%s.csv" % DEFAULT_CB_PUBLISH_FILENAME
# # ['債券簡稱', '發行人', '發行日期', '到期日期', '年期', '發行總面額', '發行資料']
# 	DEFAULT_CB_PUBLISH_FIELD_TYPE = [str, str, str, str, int, int, str,]
# 	DEFAULT_CB_PUBLISH_FIELD_TYPE_LEN = len(DEFAULT_CB_PUBLISH_FIELD_TYPE)

	SEARCH_RULE_DATASHEET_LIST = [
		["主力買超天數累計", "外資買超天數累計", "投信買超天數累計",],
		["主力買超天數累計", "外資買超天數累計",],
		["主力買超天數累計", "投信買超天數累計",],
		["主力買超天數累計",],
	]
	ETF_SEARCH_RULE_FIELD_LIST = [
		OrderedDict([("年化標準差", "低於平均"), ("Sharpe", "高於平均"), ("年報酬", "高到低排序"),]),
		OrderedDict([("Alpha", "高於平均"), ("Beta", "低於平均"), ("Sharpe", "高於平均"),]),
		OrderedDict([("年化標準差", "低於平均"), ("年報酬", "高於平均"),]),
		OrderedDict([("年化標準差", "低於平均"), ("年報酬", "高於平均"), ("Alpha", "高於平均"), ("Beta", "低於平均"), ("Sharpe", "高於平均"),]),
	]


	@classmethod
	def __is_string(cls, value):
		is_string = False
		try:
			int(value)
		except ValueError:
			is_string = True
		return is_string


	@classmethod
	def __check_file_exist(cls, filepath):
		check_exist = True
		try:
			os.stat(filepath)
		except OSError as exception:
			if exception.errno != errno.ENOENT:
				print ("%s: %s" % (errno.errorcode[exception.errno], os.strerror(exception.errno)))
				raise
			check_exist = False
		return check_exist


# ========= XLSX 工具 =========
	@classmethod
	def __read_xlsx(cls, source_filepath, sheet_name=None):
		wb = load_workbook(source_filepath)
		ws = wb.active if sheet_name is None else wb[sheet_name]
		rows = []
		header = [c.value for c in ws[1]]
		rows.append(header)
		# for row in ws.iter_rows(min_row=2, values_only=True):
		# 	rows.append(dict(zip(header, row)))
		for row in ws.iter_rows(min_row=2, values_only=True):
			rows.append(row)
		return rows


	# @classmethod
	# def __read_worksheet(cls, worksheet, sheet_metadata):
	# 	# import pdb; pdb.set_trace()
	# 	csv_data_dict = {}

	# 	sheet_name = worksheet.name
	# 	# print(sheet_name)
	# 	# if sheet_name == "法人共同買超累計":
	# 	# 	import pdb; pdb.set_trace()
	# 	start_column_index = sheet_metadata["data_start_column_index"]
	# 	title_list = ["商品",]
	# 	type_list = [str,]

	# 	sheet_ncols = sheet_metadata["sheet_columns"] if "sheet_columns" in sheet_metadata else worksheet.ncols
	# 	sheet_nrows = 9999 if "sheet_columns" in sheet_metadata else worksheet.nrows
	@classmethod
	def __read_worksheet(cls, sheet_name, worksheet_rows, sheet_metadata):
		# import pdb; pdb.set_trace()
		csv_data_dict = {}

		start_column_index = sheet_metadata["data_start_column_index"]
		title_list = ["商品",]
		type_list = [str,]

		sheet_ncols = sheet_metadata["sheet_columns"] if "sheet_columns" in sheet_metadata else len(worksheet_rows[0])
		sheet_nrows = len(worksheet_rows)
		while True:
			row_index = sheet_nrows - 1
			if worksheet_rows[row_index][0] is not None and len(str(worksheet_rows[row_index][0])) > 0:
				break
			sheet_nrows -= 1
		# import pdb; pdb.set_trace()
		for column_index in range(start_column_index, sheet_ncols):
			# title = worksheet.cell_value(0, column_index)
			# title_list.append(title)
			title_list.append(worksheet_rows[0][column_index])
		type_list.extend([int,] * (sheet_ncols - 1))

		# print("%s %d x %d" % (sheet_name, sheet_nrows, sheet_ncols))
		csv_data_value_dict = {}
		for row_index in range(1, sheet_nrows):
			data_list = []
			ignore_data = False
			stock_number = None
			product_name = None
			# key_str = worksheet.cell_value(row_index, 0)
			key_str = worksheet_rows[row_index][0]
			# print "key_str: %s" % key_str
# '''
# 				How to fix "SyntaxWarning: invalid escape sequence" in Python?
# 				The escape character (\) in Python string literals.
# 				If you want to put a literal \ in a string you may use \\:
# 				>>> print("foo \\ bar")
# 				foo \ bar
# 				Or you may use a "raw string":
# 				>>> print(r"foo \ bar")
# 				foo \ bar
# '''
			if sheet_metadata["key_mode"] == 0:
				mobj = re.match(r"([\d]{4})\s(.+)", key_str)
				if mobj is None:
					raise ValueError("%s: Incorrect format0: %s" % (sheet_name, key_str))
				stock_number = mobj.group(1)
				product_name = mobj.group(2)
			elif sheet_metadata["key_mode"] == 1:
				# mobj = re.match("([\d]{4})\.TW", key_str)
				mobj = re.match(r"([\d]{4})", str(int(key_str)))
				if mobj is None:
					raise ValueError("%s: Incorrect format1: %s" % (sheet_name, key_str))
				stock_number = mobj.group(1)
				# product_name = worksheet.cell_value(row_index, 1)
				product_name = worksheet_rows[row_index][1]
			elif sheet_metadata["key_mode"] == 2:
				mobj = re.match(r"([\d]{4})\s{2}(.+)", key_str)
				if mobj is None:
					ignore_data = True
				else:
					stock_number = mobj.group(1)
					product_name = mobj.group(2)
			elif sheet_metadata["key_mode"] == 3:
				# import pdb; pdb.set_trace()
				mobj = re.match(r"(.+)\(([\d]{4})\)", key_str)
				if mobj is None:
					# raise ValueError("%s: Incorrect format3: %s" % (sheet_name, key_str))
					ignore_data = True
				else:
					product_name = mobj.group(1)
					stock_number = mobj.group(2)
			elif sheet_metadata["key_mode"] == 4:
				# import pdb; pdb.set_trace()
				# print(f"{row_index}: {key_str}")
				if len(key_str) == 0:
					break
				# mobj = re.match("(0[\d]{3}[\dBLKRS]{0,3}) (.+)", key_str)
				mobj = re.match(r"(0[\d]{3}[\dBLKRS]{0,3})", key_str)
				if mobj is None:
					raise ValueError("%s: Incorrect format4: %s" % (sheet_name, key_str))
				stock_number = mobj.group(1)
				# product_name = worksheet.cell_value(row_index, 1)  # mobj.group(2)
				product_name = worksheet_rows[row_index][1]  # mobj.group(2)
			elif sheet_metadata["key_mode"] == 5:
				mobj = re.match("([A-Z]{2,5})", key_str)
				if mobj is None:
					raise ValueError("%s: Incorrect format5: %s" % (sheet_name, key_str))
				stock_number = mobj.group(1)
				# product_name = worksheet.cell_value(row_index, 1)
				product_name = worksheet_rows[row_index][1]
			else:
				raise ValueError("Unknown key mode: %d" % sheet_metadata["key_mode"])
			# if stock_number is None:
			#	raise RuntimeError("Fail to parse the stock number")
			if not ignore_data:
				data_list.append(product_name)
				for column_index in range(start_column_index, sheet_ncols):
					# data = worksheet.cell_value(row_index, column_index)
					data = worksheet_rows[row_index][column_index]
					if re.search("[1-9]+", str(data).split(".")[-1]) is not None:
						type_list[column_index] = float
					data_list.append(data)
			# print "%d -- %s" % (row_index, stock_number)
			value_dict = dict(zip(title_list, data_list))
			csv_data_value_dict[stock_number] = dict(zip(title_list, data_list))
		# import pdb; pdb.set_trace()
		csv_data_dict = {
			"value": csv_data_value_dict,
			"type": type_list,
		}
		# print("%s: %s" % (sheet_name, " ".join(map(lambda x, y: "%s(%s)" % (x, str(y)), title_list, type_list))))
		return csv_data_dict


	@classmethod
	def __get_file_modification_date(cls, filepath, raise_exception=True):
		if not cls.__check_file_exist(filepath):
			if raise_exception:
				raise ValueError("The file[%s] does NOT exist" % filepath)
			else:
				return None
		# print("filepath: %s" % filepath)
		# create_time = os.path.getctime(filepath)
		# print(create_time)
		modification_time = os.path.getmtime(filepath)
		# print(modification_time)
		modification_date = datetime.fromtimestamp(modification_time)
		return modification_date


	@classmethod
	def show_search_targets_list(cls):
		print("*****************************************")
		print("Targets search rule")
		for index, search_rule_dataset in enumerate(cls.SEARCH_RULE_DATASHEET_LIST):
			print(" (%d)  %s" % (index, ", ".join(search_rule_dataset)))
		print("*****************************************")
		print("ETF Targets search rule")
		for index, search_rule_dataset in enumerate(cls.ETF_SEARCH_RULE_FIELD_LIST):
			# import pdb; pdb.set_trace()
			print(" (%d)  %s" % (index, ", ".join(map(lambda x: "%s:%s" % (x[0], x[1]), search_rule_dataset.items()))))
		print("*****************************************")


	def __init__(self, cfg):
		self.xcfg = {
			"source_folderpath": None,
			"source_filename": self.DEFAULT_SOURCE_FULL_FILENAME,
			"tracked_stock_list_filename": self.DEFAULT_TRACKED_STOCK_LIST_FILENAME,
			"tracked_stock_list": None,
			"sort_tracked_stock_list_output": True,
			"min_consecutive_over_buy_days": self.DEFAULT_MIN_CONSECUTIVE_OVER_BUY_DAYS,
			"max_consecutive_over_buy_days": self.DEFAULT_MAX_CONSECUTIVE_OVER_BUY_DAYS,
			"minimum_volume": self.DEFAULT_MINIMUM_VOLUME,
			"output_result_filename": self.DEFAULT_OUTPUT_RESULT_FILENAME,
			"output_result": False,
			"quiet": False,
			"check_sharpe_ratio": False,
			"show_scrapy_progress": False,
			'exclude_sheet': None,
		}
		# import pdb; pdb.set_trace()
		self.xcfg.update(cfg)
		self.xcfg["source_folderpath"] = self.DEFAULT_SOURCE_FOLDERPATH if self.xcfg["source_folderpath"] is None else self.xcfg["source_folderpath"]
		self.xcfg["source_filename"] = self.DEFAULT_SOURCE_FULL_FILENAME if self.xcfg["source_filename"] is None else self.xcfg["source_filename"]
		self.xcfg["source_filepath"] = os.path.join(self.xcfg["source_folderpath"], self.xcfg["source_filename"])
		file_modification_date = self.__get_file_modification_date(self.xcfg["source_filepath"], raise_exception=False)
		if file_modification_date is not None:
			self.xcfg["source_file_modification_date_str"] = file_modification_date.strftime("%Y/%m/%d %H:%M:%S")
		else:
			self.xcfg["source_file_modification_date_str"] = None
		# print ("__init__: %s" % self.xcfg["source_filepath"])
		self.xcfg["tracked_stock_list_filepath"] = os.path.join(self.DEFAULT_CONFIG_FOLDERPATH, self.xcfg["tracked_stock_list_filename"])
		file_modification_date = self.__get_file_modification_date(self.xcfg["tracked_stock_list_filepath"], raise_exception=False)
		if file_modification_date is not None:
			self.xcfg["tracked_stock_list_file_modification_date_str"] = file_modification_date.strftime("%Y/%m/%d %H:%M:%S")
		else:
			self.xcfg["tracked_stock_list_file_modification_date_str"] = None
		if self.xcfg["tracked_stock_list"] is not None:
			if type(self.xcfg["tracked_stock_list"]) is str:
				tracked_stock_list = []
				for tracked_stock in self.xcfg["tracked_stock_list"].split(","):
					tracked_stock_list.append(tracked_stock)
				self.xcfg["tracked_stock_list"] = tracked_stock_list
		# import pdb; pdb.set_trace()
		self.xcfg["output_result_filepath"] = os.path.join(self.DEFAULT_CONFIG_FOLDERPATH, self.xcfg["output_result_filename"])

		self.filepath_dict = OrderedDict()
		self.filepath_dict["source"] = self.xcfg["source_filepath"]
		self.filepath_dict["tracked_stock_list"] = self.xcfg["tracked_stock_list_filepath"]
		self.filepath_dict["output_result"] = self.xcfg["output_result_filepath"]
		self.mouthly_average_return_sorted_data = None
		self.workbook = None
		self.output_result_file = None
		self.stdout_tmp = None
		self.sheet_name_list = self.DEFAULT_SHEET_NAME_LIST
		if self.xcfg['exclude_sheet'] is not None:
			self.sheet_name_list = [sheet_name for sheet_name in self.sheet_name_list if sheet_name not in self.xcfg['exclude_sheet']]
		self.__print_file_modification_date()


	def __enter__(self):
		# Open the workbook
		# self.workbook = xlrd.open_workbook(self.xcfg["source_filepath"])
		# if self.xcfg["output_result"]:
		# 	self.output_result_file = open(self.xcfg["output_result_filepath"], "w")
		return self


	def __exit__(self, type, msg, traceback):
		if self.output_result_file is not None:
			self.output_result_file.close()
			self.output_result_file = None
		if self.workbook is not None:
			self.workbook.release_resources()
			del self.workbook
			self.workbook = None
		return False


	def __get_workbook(self):
		if self.workbook is None:
			# import pdb; pdb.set_trace()
			self.workbook = xlrd.open_workbook(self.xcfg["source_filepath"])
			# print ("__get_workbook: %s" % self.xcfg["source_filepath"])
		return self.workbook


	def __redirect_stdout2file(self):
# The output is now directed to the file
		if self.output_result_file is None:
			self.output_result_file = open(self.xcfg["output_result_filepath"], 'w')
# Store the current STDOUT object for later use
		self.stdout_tmp = sys.stdout
# Redirect STDOUT to the file
		sys.stdout = self.output_result_file


	def __redirect_file2stdout(self):
# Restore the original STDOUT
		sys.stdout = self.stdout_tmp


	def __read_sheet_data(self, sheet_name):
		# import pdb; pdb.set_trace()
		sheet_metadata = self.SHEET_METADATA_DICT[sheet_name]		
# 		# print (u"Read sheet: %s" % sheet_name)
# 		# assert self.workbook is not None, "self.workbook should NOT be None"
# 		worksheet = self.__get_workbook().sheet_by_name(sheet_name)
# 		# https://www.itread01.com/content/1549650266.html
# 		# print worksheet.name,worksheet.nrows,worksheet.ncols    #Sheet1 6 4
# # The data
# 		csv_data_dict = self.__read_worksheet(worksheet, sheet_metadata)
		worksheet_rows = self.__read_xlsx(self.xcfg["source_filepath"], sheet_name=sheet_name)
		csv_data_dict = self.__read_worksheet(sheet_name, worksheet_rows, sheet_metadata)
		csv_data_value_dict = csv_data_dict["value"]
# Filter the data if necessary
		if (self.xcfg["min_consecutive_over_buy_days"] is not None) or (self.xcfg["max_consecutive_over_buy_days"] is not None):
			try:
				sheet_index = self.CONSECUTIVE_OVER_BUY_DAYS_SHEETNAME_LIST.index(sheet_name)
				# import pdb; pdb.set_trace()
				field_name = self.CONSECUTIVE_OVER_BUY_DAYS_FIELDNAME_LIST[sheet_index]
				filter_func_ptr = None
				if (self.xcfg["min_consecutive_over_buy_days"] is not None) and (self.xcfg["max_consecutive_over_buy_days"] is not None):
					filter_func_ptr = lambda x: (self.xcfg["max_consecutive_over_buy_days"] >= int(x[1][field_name]) >= self.xcfg["min_consecutive_over_buy_days"])
				elif self.xcfg["min_consecutive_over_buy_days"] is not None:
					filter_func_ptr = lambda x: (int(x[1][field_name]) >= self.xcfg["min_consecutive_over_buy_days"])
				elif self.xcfg["max_consecutive_over_buy_days"] is not None:
					filter_func_ptr = lambda x: (self.xcfg["max_consecutive_over_buy_days"] >= int(x[1][field_name]))
				if filter_func_ptr is not None:
					csv_data_value_dict = dict(filter(filter_func_ptr, csv_data_value_dict.items()))
			except ValueError as e: 
				pass
		if self.xcfg["minimum_volume"] is not None:
			try:
				sheet_index = self.MINIMUM_VOLUME_SHEETNAME_LIST.index(sheet_name)
				# import pdb; pdb.set_trace()
				field_name = self.MINIMUM_VOLUME_FIELDNAME_LIST[sheet_index]
				csv_data_value_dict = dict(filter(lambda x: int(x[1][field_name]) >= self.xcfg["minimum_volume"], csv_data_value_dict.items()))
			except ValueError as e: 
				pass
		if self.xcfg["check_sharpe_ratio"]:
			if sheet_name == self.LARGE_SHAREHOLD_POSITION_SHEETNAME:
				sharpe_ratio_sorted_list = sorted([x[self.LARGE_SHAREHOLD_POSITION_FIELDNAME_SHARPE_RATIO] for x in csv_data_value_dict.values()], reverse=True)
				standard_deviation_sorted_list = sorted([x[self.LARGE_SHAREHOLD_POSITION_FIELDNAME_STANDARD_DEVIATION] for x in csv_data_value_dict.values()])
				data_len = len(standard_deviation_sorted_list)
				for csv_data_key, csv_data_value_dict in csv_data_value_dict.items():
					sharpe_ratio_value = csv_data_value_dict[self.LARGE_SHAREHOLD_POSITION_FIELDNAME_SHARPE_RATIO]
					standard_deviation_value = csv_data_value_dict[self.LARGE_SHAREHOLD_POSITION_FIELDNAME_SHARPE_RATIO]
					sharpe_ratio_index = sharpe_ratio_sorted_list.index(sharpe_ratio_value)
					standard_deviation_index = standard_deviation_sorted_list.index(standard_deviation_value)
		csv_data_dict["value"] = csv_data_value_dict
		return csv_data_dict


	def __get_sorted_stock_list(self, sort_by_field_name, sheet_data_dict, reverse=False):
		# import pdb; pdb.set_trace()
		stock_list = [(item[0], item[1][sort_by_field_name]) for item in sheet_data_dict['value'].items()]
		sorted_stock_list = sorted(stock_list, key=lambda x: x[1], reverse=reverse)
		return sorted_stock_list


	def __get_sorted_stock_index(self, stock_id, sorted_stock_list):
		stock_order_list = [index for index, stock_data in enumerate(sorted_stock_list) if stock_data[0] == stock_id]
		if len(stock_order_list) != 1:
			raise ValueError("Incorrect search result: %s" % stock_order_list)
		return stock_order_list[0]


	def __filter_sorted_stock_list(self, sorted_stock_list, filter_percentage_threshold=50):
		sorted_stock_list_len = len(sorted_stock_list)
		filtered_list_len = int(sorted_stock_list_len * filter_percentage_threshold / 100) + 1
		return sorted_stock_list[0:filtered_list_len]


	def __get_field_in_sheet_mean(self, field_name, sheet_data_dict, reverse=False):
		# import pdb; pdb.set_trace()
		field_value_list = [item[field_name] for item in sheet_data_dict['value'].values()]
		return sum(field_value_list)/len(field_value_list)


	def __print_file_modification_date(self):
		print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
		if self.xcfg["source_file_modification_date_str"] is not None:
			print("%s  修改時間: %s" % (os.path.basename(self.xcfg["source_filepath"]), self.xcfg["source_file_modification_date_str"]))
		if self.xcfg["tracked_stock_list_file_modification_date_str"] is not None:
			print("%s  修改時間: %s" % (os.path.basename(self.xcfg["tracked_stock_list_filepath"]), self.xcfg["tracked_stock_list_file_modification_date_str"]))
		print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++\n")


	def __get_stock_chip_data(self, sheet_name_list=None):
		stock_chip_data_dict = {}
		if sheet_name_list is None:
			sheet_name_list = self.sheet_name_list
		for sheet_name in sheet_name_list:
			stock_chip_data_dict[sheet_name] = self.__read_sheet_data(sheet_name)
		# import pdb; pdb.set_trace()
		return stock_chip_data_dict


	def search_etf_targets(self, stock_chip_data_dict=None, search_rule_index=0):
		if stock_chip_data_dict is None:
			stock_chip_data_dict = self.__get_stock_chip_data()
		if search_rule_index < 0 or search_rule_index >= len(self.ETF_SEARCH_RULE_FIELD_LIST):
			raise ValueError("Unsupport ETF search_rule_index: %d" % search_rule_index)
		if self.xcfg["output_result"]:
			self.__redirect_stdout2file()
		print("************** Search ETF **************")
		search_rule_list = self.ETF_SEARCH_RULE_FIELD_LIST[search_rule_index]
		search_rule_list_str = ", ".join(map(lambda x: "%s:%s" % (x[0], x[1]), search_rule_list.items()))
		print ("搜尋規則: " + search_rule_list_str)

		# import pdb; pdb.set_trace()
		field_rule_dict = self.ETF_SEARCH_RULE_FIELD_LIST[search_rule_index]
		for sheet_name in self.ETF_SHEET_NAME_LIST:
			stock_set = None
			sheet_data_dict = stock_chip_data_dict[sheet_name]  # ["value"]
			# import pdb; pdb.set_trace()
			print("%s" % sheet_name)
			sort_field_name = None
			for field_name, field_rule in field_rule_dict.items():
				if field_rule not in ["高於平均", "低於平均",]:
					sort_field_name = None
					continue
				# reverse = False if field_rule == "低於平均" else True
				# sorted_stock_list = self.__get_sorted_stock_list(field_name, sheet_data_dict, reverse=reverse)
				# filtered_stock_list = self.__filter_sorted_stock_list(sorted_stock_list)
				mean_value = self.__get_field_in_sheet_mean(field_name, sheet_data_dict)
				filter_funcptr = (lambda x: x[1][field_name] <= mean_value) if field_rule == "低於平均" else (lambda x: x[1][field_name] >= mean_value)
				filtered_stock_list = list(filter(filter_funcptr, sheet_data_dict["value"].items()))
				filtered_stock_id_list = [filtered_stock[0] for filtered_stock in filtered_stock_list]
				if stock_set is None:
					stock_set = set(filtered_stock_id_list)
				else:
					stock_set &= set(filtered_stock_id_list)
			stock_list = list(stock_set)
			# print("%s: %s" % (sheet_name, ", ".join(stock_list)))
			# import pdb; pdb.set_trace()
			filtered_sheet_data_value_dict = dict(filter(lambda x: x[0] in stock_list, sheet_data_dict['value'].items()))
			if sort_field_name is not None:
				sorted_sheet_data_value_dict = OrderedDict(sorted(filtered_sheet_data_value_dict.items(), key=lambda x: x[1][sort_field_name], reverse=reverse))
			else:
				sorted_sheet_data_value_dict = filtered_sheet_data_value_dict
			for index, stock_data_tuple in enumerate(sorted_sheet_data_value_dict.items()):
				stock = stock_data_tuple[0]
				stock_sheet_data_dict = stock_data_tuple[1]
				stock_name = stock_sheet_data_dict["商品"]
				print ("*** %s[%s] ***" % (stock, stock_name))
				item_list = stock_sheet_data_dict.items()
				item_type_list = map(lambda x, y: (x[0], x[1], y), item_list, stock_chip_data_dict[sheet_name]["type"])
				item_type_list = filter(lambda x: x[0] not in ["商品",], item_type_list)
				try:
					print("  " + " ".join(map(lambda x: "%s(%s)" % (x[0], str(x[2](x[1]))), item_type_list)))
				except ValueError as e:
					raise e
			print("\n")


	def search_targets(self, stock_chip_data_dict=None, search_rule_index=0):
		if stock_chip_data_dict is None:
			stock_chip_data_dict = self.__get_stock_chip_data()
		if search_rule_index < 0 or search_rule_index >= len(self.SEARCH_RULE_DATASHEET_LIST):
			raise ValueError("Unsupport search_rule_index: %d" % search_rule_index)
		search_rule_list = self.SEARCH_RULE_DATASHEET_LIST[search_rule_index]
		stock_set = set(stock_chip_data_dict[search_rule_list[0]]["value"].keys())
		for search_rule in search_rule_list[1:]:
			stock_set &= set(stock_chip_data_dict[search_rule]["value"].keys())
		stock_list = list(stock_set)

		if self.xcfg["output_result"]:
			self.__redirect_stdout2file()
		print("************** Search **************")
		search_rule_list_str = ", ".join(search_rule_list)
		print ("搜尋規則: " + search_rule_list_str )
		stock_name_list = [stock_chip_data_dict[u"主力買超天數累計"]["value"][stock]["商品"] for stock in stock_list]
		stock_list_str = ", ".join(map(lambda x: "%s[%s]" % (x[0], x[1]), zip(stock_list, stock_name_list)))
		print (stock_list_str + "\n")
		# import pdb; pdb.set_trace()
		sheet_name_list = copy.deepcopy(self.SEARCH_RULE_DATASHEET_LIST[search_rule_index])
		for index, stock in enumerate(stock_list):
			print ("*** %s[%s] ***" % (stock, stock_name_list[index]))
			global_item_list = None
			for sheet_name in sheet_name_list:				
				sheet_data_dict = stock_chip_data_dict[sheet_name]["value"]
				if stock not in sheet_data_dict.keys():
					continue
				stock_sheet_data_dict = sheet_data_dict[stock]
				item_list = stock_sheet_data_dict.items()
				if global_item_list is None:
					global_item_list = []
					global_item_list.extend(filter(lambda x: x[0] in ["成交", "漲幅%", "漲跌幅",], item_list))
					global_item_list.extend(map(lambda x: (x[0], str(int(x[1]))), filter(lambda x: x[0] in ["成交量", "總量",], item_list)))
					# global_item_list.extend(search_rule_item_list)
					print(" ==>" + " ".join(map(lambda x: "%s(%s)" % (x[0], x[1]), global_item_list)))
				item_type_list = map(lambda x, y: (x[0], x[1], y), item_list, stock_chip_data_dict[sheet_name]["type"])
				item_type_list = filter(lambda x: x[0] not in ["商品", "成交", "漲幅%", "漲跌", "漲跌幅", "成交量", "總量",], item_type_list)
				try:
					print("  " + sheet_name + ": " + " ".join(map(lambda x: "%s(%s)" % (x[0], str(x[2](x[1]))), item_type_list)))
				except ValueError as e:
					raise e
			print("\n")
		if self.xcfg["output_result"]:
			self.__redirect_file2stdout()


	def track_targets(self, stock_chip_data_dict=None):
		# import pdb; pdb.set_trace()
		if self.xcfg["tracked_stock_list"] is None:
			self.__get_tracked_stock_list_from_file()
		if stock_chip_data_dict is None:
			stock_chip_data_dict = self.__get_stock_chip_data()

		if self.xcfg["output_result"]:
			self.__redirect_stdout2file()
		file_modification_date = self.__get_file_modification_date(self.xcfg["source_filepath"])
		print("檔案修改時間: %s\n" % file_modification_date.strftime("%Y/%m/%d %H:%M:%S"))
		print("************** Display **************")
		# import pdb; pdb.set_trace()
		for tracked_stock in self.xcfg["tracked_stock_list"]:
			# print ("*** %s[%s] ***" % (tracked_stock, stock_name_list[index]))
			target_caption = None
			global_item_list = None
			need_new_line = False
			for sheet_name in self.sheet_name_list:
				# print("Sheet name: %s" % sheet_name)
				sheet_data_dict = stock_chip_data_dict[sheet_name]["value"]
				if tracked_stock not in sheet_data_dict.keys():
					continue
				# print("Sheet name: %s --------> OUTPUT" % sheet_name)
				if not need_new_line:
					need_new_line = True
				stock_sheet_data_dict = sheet_data_dict[tracked_stock]
				if target_caption is None:
					target_caption = "*** %s[%s] ***" % (tracked_stock, stock_sheet_data_dict["商品"])
					print(target_caption)
				item_list = stock_sheet_data_dict.items()
				if global_item_list is None:
					global_item_list = []
					global_item_list.extend(filter(lambda x: x[0] in ["成交", "漲幅%", "漲跌幅",], item_list))
					global_item_list.extend(map(lambda x: (x[0], str(int(x[1]))), filter(lambda x: x[0] in ["成交量", "總量",], item_list)))
					print(" ==>" + " ".join(map(lambda x: "%s(%s)" % (x[0], x[1]), global_item_list)))
				item_type_list = map(lambda x, y: (x[0], x[1], y), item_list, stock_chip_data_dict[sheet_name]["type"])
				item_type_list = filter(lambda x: x[0] not in ["商品", "成交", "漲幅%", "漲跌", "漲跌幅", "成交量", "總量",], item_type_list)
				try:
					# import pdb; pdb.set_trace()
					print("  " + sheet_name + ": " + " ".join(map(lambda x: "%s(%s)" % (x[0], str(x[2](x[1]))), item_type_list)))
				# except ValueError as e:
				except Exception as e:
					print("%s:%s Error: %s in %s" % (tracked_stock, sheet_name, str(e), str(list(item_type_list))))
					import pdb; pdb.set_trace()
					raise e
			if need_new_line: 
				print("\n")
		if self.xcfg["output_result"]:
			self.__redirect_file2stdout()


	def __get_tracked_stock_list_from_file(self):
		# import pdb; pdb.set_trace()
		if not self.__check_file_exist(self.xcfg['tracked_stock_list_filepath']):
			raise RuntimeError("The file[%s] does NOT exist" % self.xcfg['tracked_stock_list_filepath'])
		self.xcfg["tracked_stock_list"] = []
		with open(self.xcfg['tracked_stock_list_filepath'], 'r') as fp:
			for line in fp:
				self.xcfg["tracked_stock_list"].extend(line.strip("\n").split(","))
				# self.xcfg["tracked_stock_list"].append(line.strip("\n"))
			if self.xcfg["sort_tracked_stock_list_output"]:
				self.xcfg["tracked_stock_list"].sort()


	def print_tracked_stock(self):
		if self.xcfg["tracked_stock_list"] is None:
			self.__get_tracked_stock_list_from_file()
		for index, tracked_stock in enumerate(self.xcfg["tracked_stock_list"]):
			print("%2d: %s" % (index + 1, tracked_stock))


	def modify_tracked_stock(self, modify_tracked_stock_list_str):
		# import pdb; pdb.set_trace()
		if self.xcfg["tracked_stock_list"] is None:
			self.__get_tracked_stock_list_from_file()
		modify_tracked_stock_list = modify_tracked_stock_list_str.split(",")
		for modify_tracked_stock in modify_tracked_stock_list:
			if modify_tracked_stock[0] == "+":
				add_tracked_stock = modify_tracked_stock[1:]
				if add_tracked_stock in self.xcfg["tracked_stock_list"]:
					print("The stock[%s] already exists in the list" % add_tracked_stock)
				else:
					self.xcfg["tracked_stock_list"].append(add_tracked_stock)
			elif modify_tracked_stock[0] == "x":
				remove_tracked_stock = modify_tracked_stock[1:]
				if remove_tracked_stock not in self.xcfg["tracked_stock_list"]:
					print("The stock[%s] does NOT exist in the list" % remove_tracked_stock)
				else:
					self.xcfg["tracked_stock_list"].remove(remove_tracked_stock)
			else:
				raise ValueError("Incorrect operator: %s" % modify_tracked_stock)
		# import pdb; pdb.set_trace()
		self.xcfg["tracked_stock_list"] = list(filter(lambda x: len(x) != 0, self.xcfg["tracked_stock_list"]))
		with open(self.xcfg['tracked_stock_list_filepath'], 'w') as fp:
			for line in self.xcfg["tracked_stock_list"]:
				fp.write("%s\n" % line)


	def search_sheets(self, search_whole=False):
		if search_whole:
			if self.xcfg['stock_list'] is not None:
				raise RuntimeError("The stock list should be None")
		else:
			if self.xcfg['stock_list'] is None:
				raise RuntimeError("The stock list should NOT be None")
			self.xcfg['stock_list'] = self.xcfg['stock_list'].split(",")
		self.__search_stock_sheets()


	def print_filepath(self):
		print("************** File Path **************")
		for key, value in self.filepath_dict.items():
			print("%s: %s" % (key, value))


if __name__ == "__main__":
# argparse 預設會把 help 文字裡的換行與多重空白「壓縮」成一行，所以你在字串裡寫的 \n 不一定會照原樣顯示。 => 建立 parser 時加上 formatter_class=argparse.RawTextHelpFormatter
	parser = argparse.ArgumentParser(description='Print help', formatter_class=argparse.RawTextHelpFormatter)
	'''
	參數基本上分兩種，一種是位置參數 (positional argument)，另一種就是選擇性參數 (optional argument)
	* example2.py
	parser.add_argument("pos1", help="positional argument 1")
	parser.add_argument("-o", "--optional-arg", help="optional argument", dest="opt", default="default")

	# python example2.py hello -o world 
	positional arg: hello
	optional arg: world
	'''
# How to add option without any argument? use action='store_true'
	'''
	'store_true' and 'store_false' - ?些是 'store_const' 分?用作存? True 和 False 值的特殊用例。
	另外，它?的默?值分?? False 和 True。例如:

	>>> parser = argparse.ArgumentParser()
	>>> parser.add_argument('--foo', action='store_true')
	>>> parser.add_argument('--bar', action='store_false')
	>>> parser.add_argument('--baz', action='store_false')
	'''
	parser.add_argument('-l', '--list_search_rule', required=False, action='store_true', help='List each search rule and exit.')
	parser.add_argument('-r', '--search_rule', required=False, help='The rule for selecting targets. Default: 0.')
	parser.add_argument('-s', '--search', required=False, action='store_true', help='Select targets based on the search rule.')
	parser.add_argument('--search_etf', required=False, action='store_true', help='Select ETF targets based on the search rule.')
	parser.add_argument('-t', '--track', required=False, action='store_true', help='Track specific targets.')
	parser.add_argument('--tracked_stock_list', required=False, help='The list of specific stock targets to be trackeded.')
	parser.add_argument('--print_filepath', required=False, action='store_true', help='Print the filepaths used in the process and exit.')
	parser.add_argument('--print_tracked_stock', required=False, action='store_true', help='Print the stock list tracked in the file and exit.')
	parser.add_argument('--modify_tracked_stock', required=False, help='The rule for selecting targets. Default: 0.')
	parser.add_argument('-o', '--output_result', required=False, action='store_true', help='Output the result to the file instead of STDOUT.')
	parser.add_argument('--output_result_filename', required=False, action='store_true', help='The filename of outputing the result to the file instead of STDOUT.')
	parser.add_argument('--exclude_sheet', required=False, help='Exclude the specific page(s) temporarily since there are something wrong with the page(s).')
	args = parser.parse_args()

	if args.list_search_rule:
		StockChipAnalysis.show_search_targets_list()
		sys.exit(0)

	cfg = {}
	if args.tracked_stock_list:
		cfg['tracked_stock_list'] = args.tracked_stock_list
	if args.output_result:
		cfg['output_result'] = True
	if args.output_result_filename:
		cfg['output_result_filename'] = args.output_result_filename
	if args.exclude_sheet:
		cfg['exclude_sheet'] = args.exclude_sheet.split(",")
	with StockChipAnalysis(cfg) as obj:
		if args.print_filepath:
			obj.print_filepath()
			sys.exit(0)
		if args.print_tracked_stock:
			obj.print_tracked_stock()
			sys.exit(0)
		if args.search:
			search_rule_index = int(args.search_rule) if args.search_rule else 0
			obj.search_targets(search_rule_index=search_rule_index)
		if args.search_etf:
			search_rule_index = int(args.search_rule) if args.search_rule else 0
			obj.search_etf_targets(search_rule_index=search_rule_index)
		if args.track:
			obj.track_targets()
		if args.modify_tracked_stock:
			obj.modify_tracked_stock(args.modify_tracked_stock)
			obj.print_tracked_stock()
